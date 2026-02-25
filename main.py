from __future__ import annotations

from typing import List, Any, Dict, Optional, Tuple
from io import BytesIO

from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, HttpUrl

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from easyfairs_widgets import (
    is_easyfairs_supported,
    get_container_id_for_url,
    fetch_easyfairs_stands_by_countries,
)

app = FastAPI(title="Fair Scraper API", version="1.2.0")


class ScrapeRequest(BaseModel):
    url: HttpUrl
    countries: List[str] = Field(default_factory=list)  # ["ES","PT"] o ["Spain","Portugal"]
    manufacturers_only: bool = True
    max_pages: int = 20
    timeout_ms: int = 25000
    deep_profile: bool = False
    debug: bool = False


class ScrapeResponse(BaseModel):
    url: str
    total: int
    results: List[Dict[str, Any]]
    meta: Dict[str, Any] = Field(default_factory=dict)


@app.get("/health")
def health():
    return {"status": "ok"}


def _normalize_countries(countries: List[str]) -> List[str]:
    if not countries:
        return ["Spain", "Portugal"]

    out: List[str] = []
    for c in countries:
        cc = (c or "").strip()
        if not cc:
            continue
        u = cc.upper()
        if u in ("ES", "ESP", "ESPAÑA", "SPAIN"):
            out.append("Spain")
        elif u in ("PT", "PRT", "PORTUGAL"):
            out.append("Portugal")
        else:
            out.append(cc)

    # dedupe
    seen = set()
    uniq = []
    for x in out:
        if x not in seen:
            seen.add(x)
            uniq.append(x)
    return uniq


def _scrape_easyfairs(req: ScrapeRequest) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    url = str(req.url)

    if not is_easyfairs_supported(url):
        raise HTTPException(
            status_code=400,
            detail=(
                "Esta URL no está en el mapping Easyfairs. "
                "Si es otra feria Easyfairs, añade el dominio->containerId en easyfairs_widgets.py "
                "EASYFAIRS_CONTAINER_MAP."
            ),
        )

    container_id = get_container_id_for_url(url)
    if not container_id:
        raise HTTPException(
            status_code=400,
            detail="URL parece Easyfairs, pero no hay containerId en el mapping. Añádelo a EASYFAIRS_CONTAINER_MAP.",
        )

    countries_norm = _normalize_countries(req.countries)

    try:
        rows, meta = fetch_easyfairs_stands_by_countries(
            event_url=url,
            container_id=container_id,
            countries=countries_norm,
            lang="es",
            query_seed="a",
            hits_per_page=100,
            timeout_s=max(5, int(req.timeout_ms / 1000)),
            max_pages=req.max_pages if req.max_pages > 0 else None,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Easyfairs widgets scrape failed: {e}")

    results: List[Dict[str, Any]] = []
    for r in rows:
        results.append(
            {
                "fabricante": r.get("name", "") or "",
                "actividad": r.get("activity", "") or "",
                "enlace_web": r.get("website", "") or "",
                "pais": r.get("country", "") or "",
            }
        )

    return results, meta


def _autosize_columns(ws: Worksheet, min_width: int = 12, max_width: int = 60) -> None:
    # Auto ancho por contenido (básico, sin depender de libs extra)
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        width = max(min_width, min(max_width, max_len + 2))
        ws.column_dimensions[col_letter].width = width


def _build_excel(results: List[Dict[str, Any]], url: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas"

    headers = ["Fabricante", "Actividad", "Enlace Web", "País"]
    ws.append(headers)

    # Estilo cabecera
    header_font = Font(bold=True)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for item in results:
        ws.append(
            [
                item.get("fabricante", ""),
                item.get("actividad", ""),
                item.get("enlace_web", ""),
                item.get("pais", ""),
            ]
        )

    # Filtros y freeze
    ws.auto_filter.ref = f"A1:D{ws.max_row}"
    ws.freeze_panes = "A2"

    _autosize_columns(ws)

    # Hoja meta (opcional)
    ws2 = wb.create_sheet("Meta")
    ws2.append(["Campo", "Valor"])
    ws2["A1"].font = header_font
    ws2["B1"].font = header_font
    ws2.append(["URL", url])
    ws2.append(["Total", str(len(results))])

    _autosize_columns(ws2, min_width=12, max_width=80)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


@app.post("/scrape_json", response_model=ScrapeResponse)
def scrape_json(req: ScrapeRequest):
    results, meta = _scrape_easyfairs(req)
    return ScrapeResponse(url=str(req.url), total=len(results), results=results, meta=meta)


@app.post("/scrape")
def scrape_excel(req: ScrapeRequest):
    # SIEMPRE devuelve Excel
    results, meta = _scrape_easyfairs(req)

    try:
        content = _build_excel(results, url=str(req.url))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel build failed: {e}")

    filename = "fabricantes_es_pt.xlsx"
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
